from __future__ import annotations

from datetime import datetime
from pathlib import Path
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
import os


class ReportExcelBuilder:
    columns = [
        "№ документа",
        "Дата регистрации",
        "Наименование документа",
        "Примечание",
        "Тип оборудования",
        "№ заводской",
        "№ заказа",
        "Маркировка",
        "№ станционный",
        "Станция / Объект",
        "Фамилия",
        "Имя",
        "Отчество",
        "Отдел",
    ]

    columns_extended = [
        "№ документа",
        "Дата регистрации",
        "Наименование документа",
        "Примечание",
        "Тип оборудования",
        "№ заводской",
        "№ заказа",
        "Маркировка",
        "№ станционный",
        "Станция / Объект",
        "Пользователь (создавший)",
    ]

    def build_report(self, rows: list[dict]) -> str:
        wb = Workbook()
        ws = wb.active
        ws.title = "Отчет"
        ws.append(self.columns)
        for r in rows:
            ws.append(
                [
                    r["doc_no"],
                    r["reg_date"],
                    r["doc_name"],
                    r["note"],
                    r["eq_type"],
                    r["factory_no"],
                    r["order_no"],
                    r["label"],
                    r["station_no"],
                    r["station_object"],
                    r["last_name"] or (r["username_fallback"] if r["username_fallback"] else ""),
                    r["first_name"] or "",
                    r["middle_name"] or "",
                    r["department"] or "",
                ]
            )
        # авто-ширина по заголовкам
        for i, col_name in enumerate(self.columns, start=1):
            ws.column_dimensions[get_column_letter(i)].width = max(len(col_name) + 2, 15)
        Path("var/exports").mkdir(parents=True, exist_ok=True)
        fname = f"var/exports/report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        wb.save(fname)
        return fname

    def build_report_extended(self, rows: list[dict]) -> str:
        """Создание расширенного отчета с колонкой пользователя"""
        wb = Workbook()
        ws = wb.active
        ws.title = "Отчет"
        
        # Заголовки
        for col, header in enumerate(self.columns_extended, 1):
            ws.cell(row=1, column=col, value=header)
        
        # Данные
        for row_idx, row_data in enumerate(rows, 2):
            for col_idx, col_name in enumerate(self.columns_extended, 1):
                if col_name == "№ документа":
                    ws.cell(row=row_idx, column=col_idx, value=row_data["doc_no"])
                elif col_name == "Дата регистрации":
                    ws.cell(row=row_idx, column=col_idx, value=row_data["reg_date"])
                elif col_name == "Наименование документа":
                    ws.cell(row=row_idx, column=col_idx, value=row_data["doc_name"])
                elif col_name == "Примечание":
                    ws.cell(row=row_idx, column=col_idx, value=row_data["note"])
                elif col_name == "Тип оборудования":
                    ws.cell(row=row_idx, column=col_idx, value=row_data["eq_type"])
                elif col_name == "№ заводской":
                    ws.cell(row=row_idx, column=col_idx, value=row_data["factory_no"])
                elif col_name == "№ заказа":
                    ws.cell(row=row_idx, column=col_idx, value=row_data["order_no"])
                elif col_name == "Маркировка":
                    ws.cell(row=row_idx, column=col_idx, value=row_data["label"])
                elif col_name == "№ станционный":
                    ws.cell(row=row_idx, column=col_idx, value=row_data["station_no"])
                elif col_name == "Станция / Объект":
                    ws.cell(row=row_idx, column=col_idx, value=row_data["station_object"])
                elif col_name == "Пользователь (создавший)":
                    ws.cell(row=row_idx, column=col_idx, value=row_data["username"])
        
        # Автоподбор ширины колонок
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Сохранение
        filename = f"var/exports/report_extended_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        os.makedirs(os.path.dirname(filename), exist_ok=True)
        wb.save(filename)
        return filename
    
    def build_report_extended_admin(self, rows: list[dict]) -> str:
        """Создание админского отчета с расширенными колонками"""
        wb = Workbook()
        ws = wb.active
        ws.title = "Админский отчет"
        
        # Заголовки для админского отчета
        admin_columns = [
            "ID документа", "№ документа", "Дата регистрации", "Наименование документа", 
            "Примечание", "Тип оборудования", "№ заводской", "№ заказа", "Маркировка",
            "№ станционный", "Станция / Объект", "Пользователь (создавший)"
        ]
        
        # Заголовки
        for col, header in enumerate(admin_columns, 1):
            ws.cell(row=1, column=col, value=header)
        
        # Данные
        for row_idx, row_data in enumerate(rows, 2):
            for col_idx, col_name in enumerate(admin_columns, 1):
                if col_name == "ID документа":
                    ws.cell(row=row_idx, column=col_idx, value=row_data["id"])
                elif col_name == "№ документа":
                    ws.cell(row=row_idx, column=col_idx, value=row_data["doc_no"])
                elif col_name == "Дата регистрации":
                    ws.cell(row=row_idx, column=col_idx, value=row_data["reg_date"])
                elif col_name == "Наименование документа":
                    ws.cell(row=row_idx, column=col_idx, value=row_data["doc_name"])
                elif col_name == "Примечание":
                    ws.cell(row=row_idx, column=col_idx, value=row_data["note"])
                elif col_name == "Тип оборудования":
                    ws.cell(row=row_idx, column=col_idx, value=row_data["eq_type"])
                elif col_name == "№ заводской":
                    ws.cell(row=row_idx, column=col_idx, value=row_data["factory_no"])
                elif col_name == "№ заказа":
                    ws.cell(row=row_idx, column=col_idx, value=row_data["order_no"])
                elif col_name == "Маркировка":
                    ws.cell(row=row_idx, column=col_idx, value=row_data["label"])
                elif col_name == "№ станционный":
                    ws.cell(row=row_idx, column=col_idx, value=row_data["station_no"])
                elif col_name == "Станция / Объект":
                    ws.cell(row=row_idx, column=col_idx, value=row_data["station_object"])
                elif col_name == "Пользователь (создавший)":
                    ws.cell(row=row_idx, column=col_idx, value=row_data["username"])
        
        # Автоподбор ширины колонок
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Сохранение
        filename = f"var/exports/report_admin_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        os.makedirs(os.path.dirname(filename), exist_ok=True)
        wb.save(filename)
        return filename
