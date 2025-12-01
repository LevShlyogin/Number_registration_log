from __future__ import annotations

from datetime import datetime

from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select
from openpyxl import load_workbook

from app.repositories.documents import DocumentsRepository
from app.repositories.equipment import EquipmentRepository
from app.repositories.users import UsersRepository
from app.repositories.doc_numbers import DocNumbersRepository
from app.repositories.counter import CounterRepository
from app.models.doc_number import DocNumStatus


class ExcelImporterService:
    def __init__(self, session: AsyncSession):
        self.session = session
        self.docs_repo = DocumentsRepository(session)
        self.eq_repo = EquipmentRepository(session)
        self.users_repo = UsersRepository(session)
        self.docnums_repo = DocNumbersRepository(session)
        self.counter_repo = CounterRepository(session)

    async def import_file(self, path: str) -> dict:
        wb = load_workbook(filename=path)
        ws = wb.active
        # Ожидаемые заголовки:
        # № документа | Дата регистрации | Наименование документа | Примечание |
        # Тип оборудования | № заводской | № заказа | Маркировка | № станционный | Станция / Объект |
        # Фамилия | Имя | Отчество | Отдел | Имя пользователя в системе
        headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        col = {h: i for i, h in enumerate(headers)}

        max_numeric = 0
        created = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            doc_no = row[col["№ документа"]]
            if not doc_no:
                continue
            numeric = int(str(doc_no).split("-")[-1])
            reg_date = row[col["Дата регистрации"]]
            if isinstance(reg_date, str):
                try:
                    reg_date = datetime.strptime(reg_date, "%d.%m.%Y")
                except Exception:
                    try:
                        reg_date = datetime.fromisoformat(reg_date)
                    except Exception:
                        reg_date = datetime.now()
            doc_name = row[col["Наименование документа"]]
            note = row[col["Примечание"]]
            eq_type = row[col["Тип оборудования"]]
            factory_no = row[col["№ заводской"]]
            order_no = row[col["№ заказа"]]
            label = row[col["Маркировка"]]
            station_no = row[col["№ станционный"]]
            station_object = row[col["Станция / Объект"]]
            last_name = row[col.get("Фамилия", None)] if "Фамилия" in col else None
            first_name = row[col.get("Имя", None)] if "Имя" in col else None
            middle_name = row[col.get("Отчество", None)] if "Отчество" in col else None
            department = row[col.get("Отдел", None)] if "Отдел" in col else None
            username = row[col.get("Имя пользователя в системе", None)] if "Имя пользователя в системе" in col else None
            username = username or "import"

            # user
            user = await self.users_repo.get_by_username(username)
            if not user:
                user = await self.users_repo.create(username)
                user.last_name = last_name
                user.first_name = first_name
                user.middle_name = middle_name
                user.department = department

            # equipment (по совокупности полей считаем "одинаковым")
            res = await self.session.execute(
                select(self.eq_repo.session.get_bind().mapper.mapped_table)
            )
            # упрощенно: создаем всегда новую запись оборудования по данным строки
            eq = await self.eq_repo.create(
                {
                    "eq_type": eq_type or "N/A",
                    "factory_no": factory_no,
                    "order_no": order_no,
                    "label": label,
                    "station_no": station_no,
                    "station_object": station_object,
                    "notes": None,
                }
            )

            # document
            doc = await self.docs_repo.create(
                {
                    "numeric": numeric,
                    "reg_date": reg_date,
                    "doc_name": doc_name or "",
                    "note": note or "",
                    "equipment_id": eq.id,
                    "user_id": user.id,
                }
            )

            # doc_numbers mark as assigned
            from app.models.doc_number import DocNumber
            dn = DocNumber(
                numeric=numeric,
                is_golden=(numeric % 100 == 0),
                status=DocNumStatus.assigned,
                reserved_by=None,
                session_id=None,
                assigned_at=reg_date,
            )
            self.session.add(dn)

            max_numeric = max(max_numeric, numeric)
            created += 1

        if max_numeric > 0:
            await self.counter_repo.set_after_import(max_numeric + 1)
        await self.session.commit()
        return {"imported": created, "next_start": max_numeric + 1 if max_numeric > 0 else 1}
